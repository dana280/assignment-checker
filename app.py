#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
בודק מטלות אקדמאיות - התנהגות ארגונית
K2P - Knowledge to People

מערכת לבדיקה אוטומטית של מטלות סטודנטים לפי מחוון 100 נקודות
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
import anthropic
import os
import re
import zipfile
import tempfile
from pathlib import Path
import docx
import PyPDF2
from io import BytesIO

# =====================================================
# הגדרות עיצוב - צבעי K2P
# =====================================================
K2P_BLUE = "#0080C8"
K2P_GREEN = "#7FBA00"
K2P_LIGHT_BLUE = "#E6F3FF"

# CSS מותאם אישית
st.markdown(f"""
<style>
    .main {{
        background-color: #F8FBFF;
    }}
    .stButton>button {{
        background-color: {K2P_BLUE};
        color: white;
        border-radius: 10px;
        padding: 10px 25px;
        font-weight: bold;
        border: none;
    }}
    .stButton>button:hover {{
        background-color: {K2P_GREEN};
    }}
    h1 {{
        color: {K2P_BLUE};
        font-weight: bold;
    }}
    h2, h3 {{
        color: {K2P_GREEN};
    }}
    .uploadedFile {{
        background-color: {K2P_LIGHT_BLUE};
        border-radius: 5px;
        padding: 10px;
    }}
    .stProgress > div > div > div > div {{
        background-color: {K2P_GREEN};
    }}
</style>
""", unsafe_allow_html=True)

# =====================================================
# פונקציות עזר
# =====================================================

def calculate_similarity(text1, text2):
    """חישוב דמיון בין שני טקסטים"""
    # המרה למילים
    words1 = set(text1.lower().split())
    words2 = set(text2.lower().split())
    
    # חישוב דמיון Jaccard
    if len(words1) == 0 or len(words2) == 0:
        return 0
    
    intersection = words1.intersection(words2)
    union = words1.union(words2)
    
    similarity = len(intersection) / len(union) * 100
    return similarity

def extract_text_from_docx(file_path):
    """חילוץ טקסט מקובץ Word"""
    try:
        doc = docx.Document(file_path)
        return '\n'.join([para.text for para in doc.paragraphs])
    except Exception as e:
        st.error(f"שגיאה בקריאת קובץ Word: {e}")
        return ""

def extract_text_from_pdf(file_path):
    """חילוץ טקסט מקובץ PDF"""
    try:
        text = ""
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        st.error(f"שגיאה בקריאת קובץ PDF: {e}")
        return ""

def extract_info_from_file(file_path, text):
    """חילוץ מידע מקובץ: שם קובץ, מספר מטלה, שם סטודנט, ת.ז"""
    filename = Path(file_path).name
    
    # חילוץ מספר מטלה מהשם
    assignment_num = ""
    match = re.search(r'מספר[_\s]*(\d+)', filename)
    if not match:
        match = re.search(r'WorkCode[_\s]*(\d+)', filename)
    if not match:
        match = re.search(r'(\d{8,9})', filename)
    
    if match:
        assignment_num = match.group(1)
    
    # חילוץ שם סטודנט
    name = ""
    name_patterns = [
        r'שם[:\s]*([א-ת\s]+)',
        r'מגיש[:\s]*([א-ת\s]+)',
        r'סטודנט[:\s]*([א-ת\s]+)',
        r'שם הסטודנט[:\s]*([א-ת\s]+)',
    ]
    
    for pattern in name_patterns:
        match = re.search(pattern, text)
        if match:
            name = match.group(1).strip()
            # ניקוי שם - רק עד סוף השם (לפני מספרים או סימנים)
            name = re.split(r'[\d\-:]+', name)[0].strip()
            break
    
    # חילוץ ת.ז
    id_num = ""
    id_patterns = [
        r'ת\.ז[:\s]*(\d{9})',
        r'ת"ז[:\s]*(\d{9})',
        r'תעודת זהות[:\s]*(\d{9})',
        r'ID[:\s]*(\d{9})',
    ]
    
    # חיפוש בתוכן ובשם הקובץ
    search_text = text + " " + filename
    for pattern in id_patterns:
        match = re.search(pattern, search_text)
        if match:
            id_num = match.group(1)
            break
    
    return filename, assignment_num, name, id_num

def check_assignment_with_claude(text, api_key):
    """בדיקת מטלה באמצעות Claude API"""
    
    client = anthropic.Anthropic(
        api_key=api_key
    )
    
    prompt = f"""את בודקת מטלה בקורס התנהגות ארגונית לפי מחוון של 100 נקודות.

**⚠️ חשוב מאוד - עקרונות כתיבת הערות:**
1. כתבי **רק** מה שחסר או מה שצריך לתקן
2. **אל תכתבי** מה שהסטודנט עשה טוב!
3. **אל תכתבי** "כתב היטב" או "מצוין" - רק חסרים!
4. אם הציון 100 - **אין הערות בכלל!**
5. סימן מינוס בסוגריים: (15-) לא (15)

**עקרון מנחה - המרצה מקלה מאוד!**
- אם הסטודנט כתב עבודה עם תוכן = 85-100!
- רק אם משהו **באמת** חסר לגמרי - תורידי נקודות
- אם כתב על התיאוריות (גם אם לא בדיוק הנכונות) - אל תורידי הרבה!
- "ניתן להרחיב" = רק 5 נק'

**מבנה המטלה:**
שאלה 1: תרבות (40 נק')
שאלה 2: 3 תיאוריות מבנה (20 נק')
שאלה 3: 2 תיאוריות תהליך (20 נק')
שאלה 4: 2 תיאוריות תוכן (20 נק')

**מחוון הבדיקה:**

**שאלה 1 - תרבות (40 נק'):**

סעיף א - תרבות כללית (15 נק'): 
- חובה: להסביר שהתרבות הכללית היא תרבות **מדינת ישראל**
- אם לא הבין שזו המדינה → "התרבות הכללית היא תרבות מדינת ישראל" (15-)
- אם כתב על תרבות כללית אחרת → "ניתן לפרט יותר על התרבות הכללית" (5-)
- אם כתב משהו על תרבות כללית → תן לפחות 10 נק'!

סעיף ב - תרבות ארגונית (15 נק'):
- אם כתב על התרבות הארגונית → 15 נק'!
- אם התיאור דל → "ניתן לפרט יותר על התרבות הארגונית" (5-)
- רק אם חסר לגמרי → "חסרה התייחסות לתרבות הארגונית" (15-)

סעיף ג - יחסי גומלין (10 נק'):
- אם כתב על הקשר → 10 נק'!
- אם דל → "ניתן להרחיב על יחסי הגומלין" (5-)
- אם חסר → "חסרה התייחסות ליחסי הגומלין" (10-)

**שאלה 2 - תיאוריות מבנה (20 נק'):**
- אם כתב על 3 תיאוריות → 20 נק'!
- אם לא בדיוק תיאוריות מבנה → "ניתן לדייק בבחירת התיאוריות" (5-)
- אם חסרה תיאוריה → "חסרה תיאוריה שלישית" (7-)

**שאלה 3 - תיאוריות תהליך (20 נק'):**
- אם כתב על 2 תיאוריות → 20 נק'!
- אם דל → "ניתן להרחיב על התיאוריות" (5-)
- אם חסרה → "חסרה תיאוריה שנייה" (10-)

**שאלה 4 - תיאוריות תוכן (20 נק'):**
- אם כתב על 2 תיאוריות → 20 נק'!
- אם דל → "ניתן להרחיב על התיאוריות" (5-)
- אם חסרה → "חסרה תיאוריה שנייה" (10-)

**⚠️ זכור - כתיבת הערות:**
- כתוב **רק** מה שחסר או צריך לשפר
- **אל תכתוב** מה שהסטודנט עשה טוב
- אם 100 - **אין הערות**
- שים סימן מינוס בסוגריים: (15-)

**פורמט התשובה:**

כתבי רק הערות על חסרים:
שאלה X: סעיף Y- [מה חסר] (מספר-)

דוגמאות:
שאלה 1: סעיף א- התרבות הכללית היא תרבות מדינת ישראל (15-)
שאלה 3: סעיף ב- ניתן להרחיב על מניע העובדים (5-)

בסוף, בשורה נפרדת, הציון:
85

**אם 100 - רק:**
100

**ללא הערות!**

תוכן המטלה:
{text}

החזר תשובה:
[הערות רק על חסרים - אם יש]
[שורה ריקה]
[ציון]
"""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        
        response_text = message.content[0].text
        
        # הפורמט החדש: הערות בשורות, שורה ריקה, ואז ציון בשורה אחרונה
        lines = response_text.strip().split('\n')
        
        # השורה האחרונה היא הציון
        grade = 0
        if lines:
            last_line = lines[-1].strip()
            # ניסיון לחלץ מספר מהשורה האחרונה
            grade_match = re.search(r'(\d+)', last_line)
            if grade_match:
                grade = int(grade_match.group(1))
        
        # כל השורות מלבד האחרונה (והריקות) הן הערות
        comments_lines = []
        for line in lines[:-1]:  # כל השורות חוץ מהאחרונה
            line = line.strip()
            if line and not line.startswith('ציון'):  # שורות לא ריקות ולא שורת ציון
                comments_lines.append(line)
        
        comments = '\n'.join(comments_lines) if comments_lines else ""
        
        return grade, comments
        
    except Exception as e:
        st.error(f"שגיאה בבדיקת המטלה: {e}")
        return 0, f"שגיאה: {str(e)}"

def create_excel_report(results_df, output_path):
    """יצירת דוח Excel מעוצב עם צבעים מתחלפים"""
    
    # המרת ת.ז לטקסט
    results_df['תז'] = results_df['תז'].astype(str).replace('nan', '').replace('', '')
    
    # שמירה בסיסית
    results_df.to_excel(output_path, index=False, engine='openpyxl')
    
    # טעינה לעיצוב
    wb = load_workbook(output_path)
    ws = wb.active
    
    # הגדרת רוחב עמודות
    ws.column_dimensions['A'].width = 35  # שם קובץ
    ws.column_dimensions['B'].width = 15  # מספר מטלה
    ws.column_dimensions['C'].width = 15  # ת.ז
    ws.column_dimensions['D'].width = 10  # ציון
    ws.column_dimensions['E'].width = 100 # הערות
    
    # צבעים בהירים מתחלפים
    colors = [
        'D6EAF8',  # תכלת בהיר
        'D5F4E6',  # ירוק בהיר
        'FCF3CF',  # צהוב בהיר
        'FADBD8',  # ורוד בהיר
    ]
    
    # עיצוב כותרות
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # עיצוב שורות נתונים עם צבעים מתחלפים
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        # בחירת צבע לפי אינדקס השורה
        color = colors[idx % len(colors)]
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
            
            # גופן מודגש לציון
            if cell.column == 4:  # עמודת הציון
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # גובה שורות אוטומטי
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = None  # גובה אוטומטי
    
    wb.save(output_path)
    return output_path

# =====================================================
# ממשק המשתמש
# =====================================================

def main():
    # כותרת עם לוגו
    col1, col2 = st.columns([1, 3])
    with col1:
        try:
            st.image("k2p_logo.png", width=200)
        except:
            pass
    with col2:
        st.title("🎓 בודק מטלות אקדמאיות")
        st.markdown(f"<h3 style='color: {K2P_GREEN};'>התנהגות ארגונית</h3>", unsafe_allow_html=True)
    
    st.markdown("---")
    
    # הגדרות API
    with st.sidebar:
        st.markdown(f"### ⚙️ הגדרות")
        
        with st.expander("🔑 הגדרות API", expanded=False):
            api_key = st.text_input(
                "Claude API Key",
                type="password",
                help="מפתח API של Anthropic"
            )
        
        st.markdown("---")
        st.markdown("### 📊 מידע")
        st.info("""
        **המערכת תבדוק:**
        - תרבות (40 נק')
        - תיאוריות מבנה (20 נק')
        - תיאוריות תהליך (20 נק')
        - תיאוריות תוכן (20 נק')
        """)
    
    # העלאת קבצים
    st.markdown("### 📁 העלאת מטלות")
    
    uploaded_files = st.file_uploader(
        "בחרו קבצי מטלות (Word/PDF) או קובץ ZIP",
        type=['docx', 'pdf', 'zip'],
        accept_multiple_files=True,
        help="ניתן להעלות עד 50 מטלות בכל פעם"
    )
    
    if uploaded_files and api_key:
        if st.button("🚀 התחל בדיקה", use_container_width=True):
            
            # יצירת תיקייה זמנית
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = Path(temp_dir)
                
                # פריסת קבצים
                files_to_check = []
                
                for uploaded_file in uploaded_files:
                    file_path = temp_path / uploaded_file.name
                    
                    with open(file_path, 'wb') as f:
                        f.write(uploaded_file.getbuffer())
                    
                    # אם ZIP - חלץ
                    if uploaded_file.name.endswith('.zip'):
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            zip_ref.extractall(temp_path / 'extracted')
                        
                        # הוסף את כל הקבצים מה-ZIP
                        for extracted_file in (temp_path / 'extracted').rglob('*'):
                            if extracted_file.is_file() and extracted_file.suffix in ['.docx', '.pdf']:
                                files_to_check.append(extracted_file)
                    else:
                        files_to_check.append(file_path)
                
                # הגבלה ל-50 מטלות
                if len(files_to_check) > 50:
                    st.warning(f"⚠️ נמצאו {len(files_to_check)} מטלות. בודק את 50 הראשונות.")
                    files_to_check = files_to_check[:50]
                
                st.success(f"✅ נמצאו {len(files_to_check)} מטלות לבדיקה")
                
                # בדיקת מטלות
                results = []
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for idx, file_path in enumerate(files_to_check):
                    status_text.text(f"בודק מטלה {idx + 1} מתוך {len(files_to_check)}: {file_path.name}")
                    
                    # חילוץ טקסט
                    if file_path.suffix == '.docx':
                        text = extract_text_from_docx(file_path)
                    else:
                        text = extract_text_from_pdf(file_path)
                    
                    if not text:
                        st.warning(f"⚠️ לא ניתן לקרוא את הקובץ: {file_path.name}")
                        continue
                    
                    # חילוץ מידע
                    filename, assignment_num, name, id_num = extract_info_from_file(file_path, text)
                    
                    # בדיקת המטלה
                    grade, comments = check_assignment_with_claude(text, api_key)
                    
                    # שמירת תוצאות (ללא שם סטודנט)
                    results.append({
                        'שם_קובץ': filename,
                        'מספר_מטלה': assignment_num,
                        'תז': id_num,
                        'ציון': grade,
                        'הערות': comments,
                        'טקסט_מלא': text  # שמירת הטקסט לבדיקת דמיון
                    })
                    
                    # עדכון progress
                    progress_bar.progress((idx + 1) / len(files_to_check))
                
                status_text.text("✅ הבדיקה הושלמה!")
                
                # בדיקת דמיון בין מטלות
                status_text.text("🔍 בודק דמיון בין מטלות...")
                for i in range(len(results)):
                    for j in range(i + 1, len(results)):
                        similarity = calculate_similarity(results[i]['טקסט_מלא'], results[j]['טקסט_מלא'])
                        
                        # אם דמיון גבוה מאוד (90%+) - סימון כזהה
                        if similarity >= 90:
                            duplicate_msg = f"⚠️ זהה למטלה מספר {results[j]['מספר_מטלה']} (דמיון {similarity:.0f}%)"
                            
                            # הוספה להערות של שתי המטלות
                            if results[i]['הערות']:
                                results[i]['הערות'] += f"; {duplicate_msg}"
                            else:
                                results[i]['הערות'] = duplicate_msg
                            
                            duplicate_msg_reverse = f"⚠️ זהה למטלה מספר {results[i]['מספר_מטלה']} (דמיון {similarity:.0f}%)"
                            if results[j]['הערות']:
                                results[j]['הערות'] += f"; {duplicate_msg_reverse}"
                            else:
                                results[j]['הערות'] = duplicate_msg_reverse
                
                # יצירת DataFrame (ללא עמודת הטקסט המלא)
                results_for_display = [{k: v for k, v in r.items() if k != 'טקסט_מלא'} for r in results]
                results_df = pd.DataFrame(results_for_display)
                
                # הצגת תוצאות
                st.markdown("### 📊 תוצאות הבדיקה")
                st.dataframe(
                    results_df,
                    use_container_width=True,
                    hide_index=True
                )
                
                # סטטיסטיקות
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("מספר מטלות", len(results_df))
                with col2:
                    st.metric("ממוצע ציונים", f"{results_df['ציון'].mean():.1f}")
                with col3:
                    st.metric("ציון מקסימלי", results_df['ציון'].max())
                with col4:
                    st.metric("ציון מינימלי", results_df['ציון'].min())
                
                # יצירת Excel להורדה
                output_buffer = BytesIO()
                output_path = temp_path / 'דוח_בדיקת_מטלות.xlsx'
                create_excel_report(results_df, output_path)
                
                with open(output_path, 'rb') as f:
                    excel_data = f.read()
                
                st.download_button(
                    label="📥 הורד דוח Excel",
                    data=excel_data,
                    file_name=f"דוח_בדיקת_מטלות_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    
    elif uploaded_files and not api_key:
        st.warning("⚠️ נא להזין Claude API Key בסרגל הצדדי")
    
    # פוטר
    st.markdown("---")
    st.markdown(
        f"<div style='text-align: center; color: {K2P_BLUE};'>"
        "מערכת בדיקת מטלות | K2P - Knowledge to People"
        "</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
